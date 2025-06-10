import openpyxl
from datetime import datetime
from typing import Dict, List
import os

class ExcelHandler:
    @staticmethod
    def create_excel_file(results: Dict[str, Dict], filename: str) -> str:
        """Создание Excel файла с результатами и статусами"""
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Email Results"
        
        # Заголовки
        headers = ["URL", "Email", "Статус", "Страница контактов", "Дата сканирования"]
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = openpyxl.styles.Font(bold=True)
        
        # Данные
        row = 2
        current_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        for url, data in results.items():
            emails = data.get("emails", [])
            status = data.get("status", "N/A")
            contact_page = data.get("contact_page", "")

            if emails:
                for email in emails:
                    sheet.cell(row=row, column=1, value=url)
                    sheet.cell(row=row, column=2, value=email)
                    sheet.cell(row=row, column=3, value=status)
                    # Делаем ссылку кликабельной
                    if contact_page:
                        sheet.cell(row=row, column=4, value=contact_page).hyperlink = contact_page
                    sheet.cell(row=row, column=5, value=current_date)
                    row += 1
            else:
                # Если email не найдены, все равно выводим строку со статусом
                sheet.cell(row=row, column=1, value=url)
                sheet.cell(row=row, column=2, value="Не найдено")
                sheet.cell(row=row, column=3, value=status)
                if contact_page:
                    sheet.cell(row=row, column=4, value=contact_page).hyperlink = contact_page
                sheet.cell(row=row, column=5, value=current_date)
                row += 1
        
        # Автоподбор ширины колонок
        for col_idx, column_cells in enumerate(sheet.columns, 1):
            max_length = 0
            # Устанавливаем минимальную ширину для заголовков
            if sheet.cell(row=1, column=col_idx).value:
                max_length = len(sheet.cell(row=1, column=col_idx).value)

            for cell in column_cells:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            # Ограничиваем максимальную ширину
            adjusted_width = min(max_length + 2, 70)
            sheet.column_dimensions[column_cells[0].column_letter].width = adjusted_width
        
        # Сохранение файла
        workbook.save(filename)
        return filename 